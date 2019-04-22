from openpyxl import Workbook
from openpyxl import load_workbook
from google.cloud import traslate

# create Workbook object
wb = Workbook()

# set file path
filepath = "C:\Users\dalla\Documents\CapstoneWork\Demo.xlsx"

# save workbook
wb.save(filepath)

# load Demo.xlsx
wb = load_workbook(filepath)
ws = wb.get_sheet_by_name('Sheet1')

# get b1 cell value (this can be any starting point in our execl sheet)
#b1 = sheet ['B1']
i = 0
for row in wb .iter_rows {'C():C()'.format(wb.min_row,wb.max_row)}
	for cell in row
		# Here is where i covert the arabic into english and store it in ExeclTran
		ExeclTran = googletranslate(B + i, "ar", "en")
		# I want to store it in the column to the right of the arabic column
		C + 'i' = ExeclTran

# Need to add another column for the english translation of Arabic text.

wb.save("Demo.xlsx")